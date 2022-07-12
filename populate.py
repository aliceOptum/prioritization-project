from xml.etree.ElementPath import prepare_descendant
from openpyxl import load_workbook
import pandas as pd
import datetime
import demo
import argparse
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

"""
def colorpicker():
    red = PatternFill(color='00008080',
                   fill_type='solid')

    green =PatternFill(start_color= '96d4a7',
                   end_color='96d4a7',
                   fill_type='solid')
    return(red)
    
"""








def getStringDate (date):
    try:
        live = str(date)
        input = datetime.datetime.strptime(live, "%Y-%m-%d")
        return(input)
    except:
        return(date)

def colorConverter (color):
    output = "NA"
    if color == "#107c1e":
        output = "On track"
    elif color == "#848689":
        output = "Not Started"
    elif color == "#21a2e0":
        output = "Complete"
    elif color == "#fce205":
        output = "At Risk"
    elif color == "#df1a7b":
        output = "Off Track"
    return output

def convertFeatures (features):
    output = "NA"
    if  "Data Migration" in features:
        output  = "DM"
    elif "Quote, Install & Bill	" in features:
        output  = "QIB"
    elif "Specialty" in features:
        output = "SP"
    elif "Consumer Engagement, Digital Therapeutic and Innovation Products" in features:
        output  = "CE DTI"
    elif "Commercial Medical Product" in features:
        output  = "CMP"
    elif "Broker and Employer Experience" in features:
        output = "BEE"
    elif "PCP - Provider Optimization (Cybertron)" in features:
        output =  "POC"
    elif "Member Experience" in features:
        output  = "ME"
    elif "Provider Experience" in features:
        output = "PE"
    elif "Benefit and Config" in features:
        output = "BC"
    elif "Data, Reporting and Analytics" in features:
        output = "RA"
    elif "Provider & Claim Payment+" in features:
        output = "PCP"
    elif "Scalability" in features:
        output  = "SC"
    elif "Digital Therapeutic and Innovation Products" in features:
        output = "DTI"
    elif  "Ops Reporting" in features:
        output = "OR"
    elif "USP" in features:
        output = "USP"
    return (output)


def colorCell (number, color):
    sheet["C" + str(number)].style = color
    sheet["D" + str(number)].style = color
    sheet["E" + str(number)].style = color
    sheet["F" + str(number)].style = color
    sheet["G" + str(number)].style = color
    sheet["H" + str(number)].style = color
    sheet["I" + str(number)].style = color
    sheet["J" + str(number)].style = color
    sheet["K" + str(number)].style = color
    sheet["L" + str(number)].style = color




parser = argparse.ArgumentParser()
parser.add_argument("predecessors", help ="enter the name of the csv file containing predecessors")
parser.add_argument("successors", help ="enter the name of the csv file containing successors")
parser.add_argument("STs", help ="enter the name of the csv file containing all of the STs")
parser.add_argument("blank", help ="enter the name of the blank excel template")
args = parser.parse_args()


demo.run(args.predecessors, args.successors, args.STs)
 

#modify this to the excel file you are writing to 
workbook = load_workbook(filename= args.blank, keep_vba= True)
workbook.iso_dates = True
 
#open workbook
sheet = workbook.active

number = 11
 
#modify this to the altered csv file name
dfData = pd.read_csv(args.STs)
#modify this to the file with the predecessors
pred  = pd.read_csv(args.predecessors)
#modify this to the file with the successors
succ = pd.read_csv(args.successors)
#modify the desired cell



for ind in dfData.index:
    value = dfData['Formatted ID'][ind]
    format_ID = "C" + str(number)
    progress = "H" + str(number)
    start_date = "I" + str(number)
    end_date = "J" + str(number)
    live1 = "L" + str(number)
    color = "G" + str(number)
    project = "B" + str(number)
    name = "D" + str(number)
    predval = "E" + str(number)
    succval = "F" + str(number)
    percentage = str(dfData['Percent Done By Story Count'][ind])

    check = colorConverter(str(dfData['Display Color'][ind]))

    if check != "Not Started":
        sheet[progress] = percentage
        sheet[format_ID] = value

        sheet[project] = convertFeatures(dfData['Project'][ind])

        sheet[start_date] = getStringDate(dfData['Planned Start Date'][ind])
        sheet[end_date] = getStringDate(dfData['Planned End Date'][ind])
        sheet[color] = colorConverter(str(dfData['Display Color'][ind]))
        sheet[name] = dfData['Name'][ind]
        sheet["Y" + str(number)] = dfData['Dependencies'][ind]

        sheet[live1] = getStringDate(dfData['Go Live Date'][ind]) 

        number += 1


    for cell in sheet['C']:
        if cell.value in pred['ID'].values:
            predval = "E"+ str(cell.row)
            index = pred[pred['ID']==cell.value]
            stvalue = index['Predecessor ID'].values[0]
            if stvalue in dfData['Formatted ID'].values:
                index = dfData.loc[dfData['Formatted ID']==stvalue].index[0]
                name = dfData['Name'][index]
                sheet[predval] = (stvalue + " " + name)

        if cell.value in succ['ID'].values:
            succval = "F" + str(cell.row)
            index = succ[succ['ID'] == cell.value]
            stvalue = (index['Successor ID'].values[0])
            if stvalue in dfData['Formatted ID'].values:
                index = dfData.loc[dfData['Formatted ID']==stvalue].index[0]
                name = dfData['Name'][index]
                sheet[succval] = (stvalue + " " + name)
"""
for cell in sheet['F']:
    next = cell.row + 1
    print(next)
    val = "C" + str(next)
    thing =  str(sheet[val].value)
    inCell = str(cell.value)
    if thing in inCell:
        color = colorpicker()
        colorCell(cell.row, color)
        colorCell (next, color )
        
"""
        

 
#modify this to file where gantt chart is displayed
workbook.save(filename= args.blank)